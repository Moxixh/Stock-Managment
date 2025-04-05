import pandas as pd
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# File reader
file_path = input('Please input the file path of the Excel file where you want the data: ')
df = pd.read_excel(file_path)
first_column_values = df.iloc[:, 0].tolist()

# Function to get Page Info
def get_page_info(symbol):
    try:
        primary_url = symbol
        page_info = requests.get(primary_url).text

        if "None" in page_info:
            fallback_url = symbol
            page_info = requests.get(fallback_url).text
        
        return page_info
    except Exception as e:
        print(f"Error fetching page for {symbol}: {str(e)}")
        return None

# Function to find P/E Ratio
def get_pe_finder(page_info, symbol):
    try:
        soup = BeautifulSoup(page_info, 'html.parser')
        pe_elements = soup.find_all('li', {'class': 'flex flex-space-between'})

        for pe_element in pe_elements:
            name_span = pe_element.find('span', class_='name')
            if name_span and 'Stock P/E' in name_span.text:
                pe_value = pe_element.find('span', class_='number').text.strip()
                print(f'Processed PE for {symbol}')
                return pe_value

        alternative_pe_element = soup.find('span', string="P/E")
        if alternative_pe_element:
            pe_value = alternative_pe_element.find_next('span', class_='number')
            return pe_value.text.strip() if pe_value else 'N/A'

        return 'P/E Not Found'
    except Exception as e:
        print(f"Error finding P/E for {symbol}: {str(e)}")
        return 'Error'

# Function to find Promoter Holding
def get_promoter_holding(page_info, symbol):
    try:
        soup = BeautifulSoup(page_info, 'html.parser')
        shareholding_section = soup.find('section', id='shareholding')
        if not shareholding_section:
            return "N/A", "N/A"

        table = shareholding_section.find('table', {'class': 'data-table'})
        if not table:
            return "N/A", "N/A"

        row = table.find('tr', {'class': 'stripe'})
        if not row:
            return "N/A", "N/A"

        first_td = row.find('td', class_='text')
        if first_td:
            button = first_td.find('button')
            if button and 'Promoters' in button.text:
                td_elements = row.find_all('td')
                percentages = [td.text.strip() for td in td_elements[1:]]
                last_value = percentages[-1] if len(percentages) >= 1 else "N/A"
                fifth_last_value = percentages[-5] if len(percentages) >= 5 else "N/A"
                print(f'Processed Promoter Holdings for {symbol}')
                return last_value, fifth_last_value
        return "N/A", "N/A"
    except Exception as e:
        print(f"Error extracting promoter holdings for {symbol}: {str(e)}")
        return "N/A", "N/A"

# Function to get EPS value
def get_eps_value(page_info, symbol):
    try:
        soup = BeautifulSoup(page_info, 'html.parser')
        profit_loss_section = soup.find('section', {'id': 'profit-loss'})
        if not profit_loss_section:
            return "N/A"

        table = profit_loss_section.find('table', {'class': 'data-table responsive-text-nowrap'})
        if not table:
            return "N/A"

        rows = table.find_all('tr')
        for row in rows:
            header = row.find('td', class_='text')
            if header and 'EPS in Rs' in header.text:
                cells = row.find_all('td')
                if cells:
                    print(f'Processed EPS for {symbol}')
                    return cells[-1].text.strip()
        return "N/A"
    except Exception as e:
        print(f"Error finding EPS for {symbol}: {str(e)}")
        return "Error"

# Function to get Median PE from chart using Selenium
def get_median_pe_from_chart(symbol):
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    try:
        url = symbol
        driver.get(url)
        time.sleep(2)

        pe_button = driver.find_element(By.XPATH, "//button[contains(text(), 'PE Ratio')]")
        pe_button.click()
        time.sleep(2)

        chart_legend = driver.find_element(By.XPATH, "//div[@id='chart-legend']")
        median_pe_element = chart_legend.find_element(By.XPATH, "//span[contains(text(), 'Median PE')]")
        median_pe = median_pe_element.text.split('=')[-1].strip()
        print(f'Processed Median PE for {symbol}')
        return median_pe
    except Exception as e:
        print(f"Error getting median P/E for {symbol}: {str(e)}")
        return "Error"
    finally:
        driver.quit()

# Function to get Sector
def get_sector(page_info, symbol):
    try:
        # Parse the HTML content using BeautifulSoup
        soup = BeautifulSoup(page_info, 'html.parser')
        
        # Locate the section containing sector information
        peers_section = soup.find('section', id='peers')
        if not peers_section:
            print(f"No peers section found for {symbol}")
            return 'Sector Not Found'
        
        # Attempt to find the paragraph element with the sector link
        sector_paragraph = peers_section.find('p', class_='sub')
        if not sector_paragraph:
            print(f"No sector paragraph found for {symbol}")
            return 'Sector Not Found'
        
        sector_link = sector_paragraph.find('a')
        if not sector_link or not sector_link.text.strip():
            print(f"No sector link found for {symbol}")
            return 'Sector Not Found'
        
        # Extract and return the text of the sector link
        sector = sector_link.text.strip()
        print(f"Processed Sector for {symbol}: {sector}")
        return sector

    except Exception as e:
        print(f"Error extracting sector for {symbol}: {str(e)}")
        return 'Error'

# Main loop to process symbols
pe_values = []
current_quarter_values = []
quarter_before_values = []
eps_values = []
median_pe_values = []
sector_values = []
discount_premium_values = []

for symbol in first_column_values:
    page_info = get_page_info(symbol)

    stock_pe = get_pe_finder(page_info, symbol)
    last_value, fifth_last_value = get_promoter_holding(page_info, symbol)
    eps_value = get_eps_value(page_info, symbol)
    median_pe_value = get_median_pe_from_chart(symbol)
    sector = get_sector(page_info, symbol)

    pe_values.append(stock_pe)
    current_quarter_values.append(last_value)
    quarter_before_values.append(fifth_last_value)
    eps_values.append(eps_value)
    median_pe_values.append(median_pe_value)
    sector_values.append(sector)

    # Calculate Stock Discount/Premium
    if median_pe_value != "Error" and median_pe_value != "N/A" and stock_pe != "Error" and stock_pe != "N/A":
        try:
            stock_pe_float = float(stock_pe.strip('%')) if '%' in stock_pe else float(stock_pe)
            median_pe_float = float(median_pe_value.strip('%')) if '%' in median_pe_value else float(median_pe_value)
            discount_premium = ((stock_pe_float - median_pe_float) / median_pe_float) * 100
            discount_premium_values.append(f"{discount_premium:.2f}%")
        except ValueError:
            discount_premium_values.append("N/A")
    else:
        discount_premium_values.append("N/A")

    time.sleep(1)

# Update DataFrame and save
df['Stock Current P/E'] = pe_values
df['Median P/E'] = median_pe_values
df['Current Quarter - Promoter Holding'] = current_quarter_values
df['4 Quarter before Promoter Holding'] = quarter_before_values
df['EPS'] = eps_values
df['Sector'] = sector_values
df['Stock Discount/Premium'] = discount_premium_values

# Save DataFrame to Excel
df.to_excel(file_path, index=False)

wb = load_workbook(file_path)
ws = wb.active

# Find the Stock Discount/Premium
discount_premium_column = None
for col_num, col_name in enumerate(df.columns, 1):
    if col_name == "Stock Discount/Premium":
        discount_premium_column = col_num
        break

if discount_premium_column is not None:
    # Define color fills
    dark_green_fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
    light_green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    brown_fill = PatternFill(start_color="A52A2A", end_color="A52A2A", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Apply color to each row in the Stock Discount/Premium column
    for row in range(2, len(discount_premium_values) + 2): 
        cell = ws.cell(row=row, column=discount_premium_column)
        value = discount_premium_values[row - 2]

        if value != "N/A" and value != "Error":
            try:
                value = float(value.strip('%'))  # Convert to float after stripping percent symbol
                # Determine the color based on the value
                if value <= -40:
                    cell.fill = dark_green_fill
                elif -40 < value <= -20:
                    cell.fill = light_green_fill
                elif -20 < value < 0:
                    cell.fill = yellow_fill
                elif value > 50:
                    cell.fill = red_fill
                else:
                    cell.fill = brown_fill
            except ValueError:
                pass

    # Save the workbook with updated colors
    wb.save(file_path)
    print("Successfully Updated with color coding.")
else:
    print("Stock Discount/Premium column not found, but other changes done.")
